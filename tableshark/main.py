# -*- coding: utf-8 -*-
from __future__ import annotations
import re
from typing import Generator, Sized, Iterable, Any, Callable, Literal, Iterator, List, overload, AnyStr
from abc import ABC, abstractmethod
import pandas as pd
import xml.etree.ElementTree as ET
import csv
import openpyxl
import io
from itertools import zip_longest
from uuid import uuid4
import json
import math
from dataclasses import dataclass, field

try:
    from tools import multiple_replace
except ImportError:
    def multiple_replace(text: str, dict_of_replaces: dict[str]) -> str:
        """
        Replaces multiple patterns in a string with their corresponding values from a dictionary.

        Args:
            text (str): The input string to perform replacements on.
            dict_of_replaces (dict[str]): A dictionary mapping patterns to their replacement values.

        Returns:
            str: The modified string with patterns replaced.

        https://stackoverflow.com/questions/15175142/how-can-i-do-multiple-substitutions-using-regex
        """
        # Create a regular expression from the dictionary keys
        regex = re.compile("(%s)" % "|".join(map(re.escape, dict_of_replaces.keys())))
        # For each match, look-up corresponding value in dictionary
        return regex.sub(lambda mo: dict_of_replaces[mo.string[mo.start():mo.end()]], text)

TableError = type('TableError', (Exception,), {})


def _define_header(text: str, structure: Literal['txt', 'xml'] = 'txt') -> str:
    """
    Defines the header section of the text.

    Args:
        text (str): The input text.

    Returns:
        str: The header section of the text.

    Raises:
        None.
    """
    if structure == 'txt':
        i = 1
        split = text.split('\n')
        find = True
        while find:
            if split[i].startswith('+--'):
                find = False
            i += 1
        return '\n'.join(split[:i])
    else:
        return text


def _define_body(text: str, structure: Literal['txt', 'xml'] = 'txt') -> str:
    """
    Defines the body section of the text.

    Args:
        text (str): The input text.

    Returns:
        str: The body section of the text.

    Raises:
        None.
    """
    if structure == 'txt':
        i = 1
        split = text.split('\n')
        find = True
        while find:
            if split[i].startswith('+--'):
                find = False
            i += 1
        return '\n'.join(split[i - 1:])
    else:
        return text


def _prepare_table(text: str, structure: Literal['txt', 'xml'] = 'txt') -> str:
    """
    Prepares a table in the given text for better readability.

    Args:
        text (str): The input text containing a table.

    Returns:
        str: The modified text with improved table formatting.

    Raises:
        None.
    """
    if structure == 'txt':
        readable = text.strip()
        readable = multiple_replace(
            readable,
            {
                '|--': '+--',
                '--|': '--+',
                '--|--': '--+--',
                '||': '|'
            }
        )
    else:
        readable = text.replace('&#x', 'bad_string')
    return readable


def _is_table(text: str) -> bool:
    """
    Checks if the given text contains a table.

    Args:
        text (str): The input text.

    Returns:
        bool: True if a table is found, False otherwise.

    Raises:
        None.
    """
    return bool(re.search(r'[+][-+]+[+]\n', text))


class Header:
    """
    Represents the header section of a table.

    The Header class is responsible for parsing and storing the columns of the header in a table.
    It provides methods for retrieving, setting, and iterating over the columns.

    Attributes:
        raw_header (str): The raw header string.
        _columns (list[str]): The parsed columns of the header.

    Methods:
        - parse_header(raw_header: str, structure: Literal['xml', 'txt']) -> list[str]:
            Parses the raw header based on the specified structure type and returns the parsed columns.

        - columns() -> list:
            Returns the columns of the header.

        - columns(columns: list):
            Sets the columns of the header.

        - rows():
            Alias for the `columns()` method.

        - __len__() -> int:
            Returns the number of columns in the header.

        - __iter__() -> Iterator[str]:
            Returns an iterator over the columns of the header.

        - __getitem__(item: Union[str, int, Tuple[int], Tuple[int, int]]) -> Union[str, List[str]]:
            Retrieves a column based on its name or index.
            If `item` is a string, it returns the index of the column with the specified name.
            If `item` is an integer, it returns the column at the specified index.
            If `item` is a tuple of integers, it returns the column at the specified 2-dimensional coordinates.
            If `item` is a slice object, it returns a sublist of columns based on the slice.

        - __setitem__(key: Union[str, int], value: str):
            Sets the value of a column based on its name or index.

        - __repr__() -> str:
            Returns a string representation of the header.

        - __add__(other: Union[str, List[str], Tuple[str], Header]) -> Header:
            Concatenates the header with another header or a column.
            Returns a new Header object with the concatenated columns.

        - __iadd__(other: Union[str, List[str], Tuple[str], Header]):
            Concatenates the header with another header or a column in-place.
    """

    def __init__(self, *columns: AnyStr, raw_header: str = '', structure: Literal['xml', 'txt'] = 'txt'):
        """
        Represents the header section of a table.

        Args:
            raw_header (str, optional): The raw header string. Defaults to ''.
            structure (Literal['xml', 'txt'], optional): The structure type of the table. Defaults to 'txt'.

        Attributes:
            raw_header (str): The raw header string.
            _columns (list[str]): The parsed columns of the header.

        Raises:
            TableError: If the structure type is unknown.
        """
        self.raw_header = raw_header
        if raw_header != '':
            self._columns = self.parse_header(raw_header, structure)
        else:
            self._columns = list(columns)

    def parse_header(self, raw_header: str, structure: Literal['xml', 'txt']) -> list[str]:
        """
        (\s*[а-яА-Я \w\d/.,()]+\s*) block pattern
        """
        if structure == 'txt':
            return self._parse_header_txt(raw_header)
        elif structure == 'xml':
            return self._parse_header_xml(raw_header)
        else:
            raise TableError('Unknown structure type')

    def _parse_header_xml(self, raw_header: str) -> list[str]:
        dict_header = ET.fromstring(raw_header)
        return list(max([[key for key in line.keys()] for line in dict_header], key=len))

    def _parse_header_txt(self, raw_header: str) -> list[str]:
        """
        (\s*[а-яА-Я \w\d/.,()]+\s*) block pattern
        """
        header = multiple_replace(raw_header, {
            ' +-': ' |-',
            '-+ ': '-| ',
            '-+-': '-|-',
            '-+\n': '-|\n'
        })
        pattern_first_sub = re.compile(r'\b-\b|\b-\B')  # escaping line break "wo-\nrd"
        header = re.sub(pattern_first_sub, '~', header)
        pattern_sub = re.compile(r'-')
        header = re.sub(pattern_sub, ' ', header)
        split_header = header.split('\n')
        number_of_columns = max([line.count('|') for line in split_header[1:-1]]) - 1
        map_blocks = [[len(block) for block in split.split('|')[1:-1]] for split in header.split('\n')[1:-1]]
        repeating_map = self._repeating_map(map_blocks)
        blocks: list[list[str]] = [[] for _ in range(number_of_columns)]
        l_p = 0  # last position
        i = 0
        for line_block, line, repeat_row_counter in zip(map_blocks, split_header[1:-1], repeating_map):
            for length_of_block, repeat_number in zip(line_block, repeat_row_counter):
                l_p += 1
                block = line[l_p:length_of_block + l_p]
                blocks[i].append(block)
                l_p += length_of_block
                while repeat_number > 0:
                    i += 1
                    blocks[i].append(block)
                    repeat_number -= 1
                i += 1
                if i % number_of_columns == 0:
                    i = i % number_of_columns
            l_p = 0
        parsed_header = self._build_header(blocks)
        return parsed_header

    @staticmethod
    def _repeating_map(map_blocks: list[list[int]]) -> list[list[int]]:
        out = [[] for _ in map_blocks]
        for i, row in enumerate(map_blocks):
            if i == len(map_blocks) - 1:
                out[i].extend([0 for _ in range(len(map_blocks[-1]))])
                break  # stop if it is last row
            next_row = map_blocks[-1]
            # next_row_lim = len(next_row)
            pos = 0
            for r1 in row:
                r2 = next_row[pos]
                n = 0
                if r1 == r2:
                    out[i].append(n)
                else:
                    while r1 != r2:
                        n += 1
                        r2 += next_row[pos + 1] + 1
                        pos += 1
                    out[i].append(n)
                pos += 1
        return out

    @staticmethod
    def _build_header(blocks: list[list[str]]) -> list[str]:
        # todo add check for len and number of columns: logN(len) is integer
        pattern_double_spaces = re.compile(r' {2,}')

        def concat(lines: list[list[str]]) -> Generator[str]:
            for line in lines:
                new_word = ""
                for word in line:
                    new_word += " " + word.strip()
                new_word = re.sub(pattern_double_spaces, ' ', new_word)
                yield new_word.replace('~ ', '').strip()

        out = []
        for final in concat(blocks):
            out.append(final)
        return out

    @property
    def columns(self) -> list:
        return self._columns

    @columns.setter
    def columns(self, value: list):
        self._columns = value

    @property
    def rows(self):
        return self._columns

    def __len__(self):
        return len(self._columns)

    def __iter__(self):
        for column in self._columns:
            yield column

    def __getitem__(self, item: str | int | tuple[int] | tuple[int, int]):
        if isinstance(item, str):
            if item in self._columns:
                return self._columns.index(item)
            else:
                raise KeyError(f'Column "{item}" does not exist!')
        elif isinstance(item, int):
            if item <= len(self) or item < 1:
                return self._columns[item]
            else:
                raise KeyError(f'This header has only {len(self)} columns, not {item}!')
        elif isinstance(item, tuple):
            if len(item) == 2:
                return self.__getitem__(item[1])
            else:
                raise KeyError(f'Can get only 2-dimensional coordinates.')
        elif isinstance(item, slice):
            return self._columns[item]

    def __setitem__(self, key: str | int, value: str):
        if isinstance(key, int):
            self._columns[key] = value
        elif isinstance(key, str):
            if key in self._columns:
                self.__setitem__(self._columns.index(key), value)
            else:
                raise KeyError(f'Column "{key}" does not exist!')

    def __repr__(self) -> str:
        return f"Header[{', '.join([str(_) for _ in self])}]"

    def __add__(self, other):
        if isinstance(other, Cell):
            header = Header()
            header._columns = self._columns + [other.value]
            return header
        if isinstance(other, str):
            header = Header()
            header._columns = self._columns + [other]
            return header
        elif isinstance(other, list) or isinstance(other, tuple) or isinstance(other, Header):
            header = Header()
            header._columns = self._columns + list(other)
            return header

    def __iadd__(self, other):
        if isinstance(other, str):
            self._right_add(other)
        elif isinstance(other, list) or isinstance(other, tuple) or isinstance(other, Header):
            for _ in other:
                self._right_add(str(_))

    def _right_add(self, other: str):
        self._columns.append(other)


@dataclass
class Cell:
    def __init__(self, value: Any):
        if hasattr(value, 'value'):
            self._value = value.value
        else:
            self._value = value
        self._type = type(value)
        self._name = ''

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, new_name: str):
        self._name = new_name

    @property
    def value(self):
        return self._value

    @value.setter
    def value(self, new_value):
        self._value = new_value

    # @property
    # def type(self):
    #     return self._type

    # def new_type(self, new: Callable):
    #     self._value = new(self._value)
    #     self._type = type(self._value)

    def __str__(self):
        return f"{self._value}"

    def __repr__(self):
        return f"Cell({self._value})"

    def __len__(self):
        return len(self._value)

    def __eq__(self, other):
        return self._value == other

    def __ne__(self, other):
        return self._value != other

    def __lt__(self, other):
        return self._value < other

    def __le__(self, other):
        return self._value <= other

    def __gt__(self, other):
        return self._value > other

    def __ge__(self, other):
        return self._value >= other

    def __add__(self, other):
        return self._value + other

    def __sub__(self, other):
        return self._value - other

    def __mul__(self, other):
        return self._value * other

    def __truediv__(self, other):
        return self._value / other

    def __floordiv__(self, other):
        return self._value // other

    def __mod__(self, other):
        return self._value % other

    def __pow__(self, other):
        return self._value ** other

    def __lshift__(self, other):
        return self._value << other

    def __rshift__(self, other):
        return self._value >> other

    def __and__(self, other):
        return self._value & other

    def __xor__(self, other):
        return self._value ^ other

    def __or__(self, other):
        return self._value | other

    def __iadd__(self, other):
        self._value += other
        return self

    def __isub__(self, other):
        self._value -= other
        return self

    def __imul__(self, other):
        self._value *= other
        return self

    def __itruediv__(self, other):
        self._value /= other
        return self

    def __ifloordiv__(self, other):
        self._value //= other
        return self

    def __imod__(self, other):
        self._value %= other
        return self

    def __ipow__(self, other):
        self._value **= other
        return self

    def __ilshift__(self, other):
        self._value <<= other
        return self

    def __irshift__(self, other):
        self._value >>= other
        return self

    def __iand__(self, other):
        self._value &= other
        return self

    def __ixor__(self, other):
        self._value ^= other
        return self

    def __ior__(self, other):
        self._value |= other
        return self

    def __neg__(self):
        return -self._value

    def __pos__(self):
        return +self._value

    def __abs__(self):
        return abs(self._value)

    def __invert__(self):
        return ~self._value

    def __round__(self, n=None):
        return round(self._value, n)

    def __floor__(self):
        return math.floor(self._value)

    def __ceil__(self):
        return math.ceil(self._value)

    def __trunc__(self):
        return math.trunc(self._value)

    def __index__(self):
        return self._value.__index__()

    def __format__(self, format_spec):
        return format(self._value, format_spec)

    def __hash__(self):
        return hash(self._value)

    def __bool__(self):
        return bool(self._value)

    def __getattr__(self, key):
        return getattr(self._value, key)

    def __setattr__(self, key, value):
        object.__setattr__(self, key, value)


class Vector(ABC):
    # _cells: list[Cell] = field(default_factory=list)
    def __init__(self, *cells: Cell | Any):
        self._cells = []
        for i, c in enumerate(cells):
            # setattr(self, str(i), c)
            if not isinstance(c, Cell):
                c = Cell(c)
            self._cells.append(c)

    def __setattr__(self, key, value):
        object.__setattr__(self, key, value)

    def __len__(self):
        return len(self._cells)

    def __getitem__(self, key: str | int) -> Any:
        """
        Get the value of a cell in the vector by cell name or index.

        Args:
            key (str | int): Cell name or index of the element.

        Returns:
            Any: The value of the cell.

        Raises:
            KeyError: If the cell name or index is not found in the vector.

        """
        if isinstance(key, str):
            for cell in self._cells:
                if cell.name == key:
                    return cell.value
            raise KeyError(f"Cell name '{key}' not found in the vector.")
        elif isinstance(key, int):
            if key < len(self._cells):
                return self._cells[key].value
            raise KeyError(f"Index '{key}' out of range.")
        else:
            raise TypeError("Invalid key type. Expected str or int.")

    def __setitem__(self, key, value):
        if isinstance(key, int):
            if not isinstance(value, Cell):
                value = Cell(value)
            self._cells[key] = value
        elif isinstance(key, str):
            for cell in self._cells:
                if cell.name == key:
                    cell.value = value
                    return
            raise KeyError(f"Cell name '{key}' not found in the vector.")
        else:
            raise TypeError("Invalid key type. Expected str or int.")

    def set_keys(self, keys: Iterable):
        for i, k in enumerate(keys):
            self._cells[i].name = k

    def get(self, value: str | int, default: Any = None) -> Any:  # todo value should be any
        """
        Get the value of a cell in the vector by cell name or index.

        Args:
            value (Any): Cell name or index of the element.
            default (Any): Default value if the cell is not found.

        Returns:
            Any: The value of the cell.

        """
        try:
            return self[value]
        except KeyError:
            return default

    def cell(self, i) -> Cell:
        return self[i]

    @property
    def cells(self) -> list[Cell]:
        return self._cells

    def to_dict(self, keys: Iterable = '') -> dict:
        # todo add keys
        pairs = []
        if keys:
            for k, v in zip_longest(keys, self._cells, fillvalue=uuid4()):
                pairs.append((k, v.value))
        else:
            for i, c in enumerate(self._cells):
                if c.name:
                    pairs.append((c.name, c.value))
                else:
                    pairs.append((str(i), c.value))
        return dict(pairs)

    def to_list(self) -> list:
        return [c.value for c in self._cells]

    def to_tuple(self) -> tuple:
        return tuple(self.to_list())

    def to_set(self) -> set:
        return set(self.to_list())

    def to_series(self) -> pd.Series:
        return pd.Series(self.to_list())

    def _repr_html_(self) -> str:
        return self.df()._repr_html_()

    def dtypes(self):
        # return a vector data type
        return self.df().dtypes

    def __eq__(self, other):
        if isinstance(other, Vector):
            if len(self) == len(other):
                for i, cell in enumerate(self):
                    if cell != other[i]:
                        return False
                return True
            else:
                return False
        else:
            return False

    def __iter__(self) -> Iterator[Cell]:
        for cell in self._cells:
            yield cell

    def __repr__(self):
        return f"{self.__class__.__name__}{self._cells}"

    def __str__(self):
        return self._str()

    def __contains__(self, item):
        return Cell(item) in self._cells

    @abstractmethod
    def df(self) -> pd.DataFrame:
        pass

    @abstractmethod
    def _str(self) -> str:
        pass

    def __iadd__(self, other):
        if isinstance(other, Cell):
            self._cells.append(other)
        elif isinstance(other, Vector):
            for cell in other:
                self._cells.append(cell)
        return self

    def __add__(self, other: Cell | Vector):
        if isinstance(other, Cell):
            return self.__add__(Vector(other))
        elif isinstance(other, Vector):
            for i, cell in zip(range(len(self), len(other) + 1), other):
                # setattr(self, str(i), cell)
                self._cells.append(cell)
        return Vector(*self._cells)

    def from_raw(self, value: Iterable[Any]):
        for i, val in enumerate(value):
            c = Cell(val)
            self._cells += (c,)
            setattr(self, str(i), c)
        return self

    def apply(self, func: Callable[[Cell, Any], Cell], *args: Any) -> 'Vector':
        """
        Apply a function to each cell in the vector, with additional arguments.

        Args:
            func (Callable[[Cell, Any], Cell]): The function to apply to each cell.
            *args (Any): Additional arguments to pass to the function.

        Returns:
            Vector: A new vector with the applied function.

        """
        new_cells = [func(cell, *args) for cell in self._cells]
        return self.__class__(*new_cells)


class Row(Vector):
    def __init__(self, *cells: Cell | Any):
        super().__init__(*cells)

    def _str(self):
        return 'Row[' + ' '.join([str(val) for val in self]) + ']'

    def df(self) -> pd.DataFrame:
        _dict = {i: val.value for i, val in enumerate(self)}
        return pd.DataFrame(_dict, index=[0])

    def _T(self):
        return Column(*self)


class Column(Vector):
    def __init__(self, *cells: Cell | Any):
        super().__init__(*cells)

    def _str(self):
        return '\n'.join([f'{i} {val}' for i, val in enumerate(self)])

    def df(self) -> pd.DataFrame:
        return pd.DataFrame({0: [val.value for val in self]})

    def _T(self):
        return Row(*self)


class Body:
    """
    Represents the body of a table.

    Args:
        raw_body (str, optional): The raw body content of the table. Default is an empty string.

    Attributes:
        raw_body (str): The raw body content of the table.
        rows (list[Row]): The list of rows in the body.
        number_of_columns (int): The number of columns in the body.
        columns (list[Column]): The list of columns in the body.

    Methods:
        parse_body(raw_body: str = '') -> list[Row]:
            Parse the raw body content and return a list of rows.

    Private Methods:
        _define_number_of_columns() -> int:
            Determine the number of columns in the body.

        _normalize_row(row: str, number_of_columns: int) -> list[Cell]:
            Normalize a row by splitting it into cells.
    """

    def __init__(self, *rows: Iterable | Row, raw_body: str = '', structure: Literal['xml', 'txt'] = 'txt'):
        """
        Initialize the Body instance.

        Args:
            raw_body (str, optional): The raw body content of the table. Default is an empty string.
        """
        self.raw_body = raw_body
        self.structure = structure
        # if not rows:
        #     rows = []
        self._rows = []
        for row in rows:
            if not isinstance(row, Row):
                self._rows.append(Row(*row))
            else:
                self._rows.append(row)
        # self._rows = rows
        if self.raw_body != '':
            self._rows: list[Row] = self.parse_body(self.raw_body, structure)
        self._number_of_columns = self._define_number_of_columns()
        self._columns: list[Column] = self._prepare_columns()

    def parse_body(self, raw_body: str = '', structure: Literal['xml', 'txt'] = 'txt') -> list[Row]:
        if structure == 'txt':
            return self._parse_body_txt(raw_body)
        elif structure == 'xml':
            return self._parse_body_xml(raw_body)
        else:
            raise TableError('Unknown structure type')

    def _parse_body_xml(self, raw_body: str) -> list[Row]:
        pass

    def _parse_body_txt(self, raw_body: str) -> list[Row]:
        body = raw_body
        pattern_first_sub = re.compile(r'\b-\b|\b-\B')  # escaping line break "wo-\nrd"
        body = re.sub(pattern_first_sub, '~', body)
        number_of_columns = self._define_number_of_columns()
        split_body_modified = re.sub(r'\+[-+]+\+', '{newline}', body).split('{newline}')
        rows = []
        for row_dirty in split_body_modified[1:-1]:
            rows.append(Row(*self._normalize_row(row_dirty, number_of_columns)))
        return rows

    def _define_number_of_columns(self) -> int:
        """
        Determine the number of columns in the body.
        :return: The number of columns in the body.
        """
        if self._rows:
            return len(self._rows[0])
        if self.structure == 'txt':
            if self.raw_body == '':
                return 0
            return max([row.count('|') for row in self.raw_body.split('\n')]) - 1
        elif self.structure == 'xml':
            return 0
        else:
            return 0

    @staticmethod
    def _normalize_row(row: str, number_of_columns: int) -> list[Cell]:
        future_row = [[] for _ in range(number_of_columns)]
        pattern_block = re.compile(r'\s*[а-яА-Я \w\d/.,():~-]+\s*')
        for sub_row in row.split('|\n')[:-1]:
            for column_number, block in zip(range(number_of_columns), re.finditer(pattern_block, sub_row)):
                future_row[column_number].append(block.group().strip())

        def prepare_cell(list_strings: list[str]) -> Cell:
            string = ' '.join(list_strings)
            string = re.sub(r'~ *', '', string).strip()
            return Cell(string)

        return list(map(prepare_cell, future_row))

    def __iter__(self) -> Iterator[Row]:
        for row in self._rows:
            yield row

    @property
    def number_of_columns(self):
        return self._number_of_columns

    @property
    def rows(self) -> list[Row]:
        return self._rows

    @rows.setter
    def rows(self, value):
        self._rows = value

    @property
    def columns(self):
        return self._columns

    def _prepare_columns(self):
        columns = []
        for col in zip(*self):
            columns.append(Column(*col))
        return columns

    def __len__(self):
        return len(self._rows)

    def __repr__(self):
        return f'Body {len(self)} rows {self._number_of_columns} columns'

    def __getitem__(self, item: int):
        return self._rows[item]

    def __add__(self, other):
        if isinstance(other, Row):
            l_ = self._rows + [other]
            return Body(*l_)
        if isinstance(other, Column):
            new_rows = []
            for my_row, value in zip(self, other):
                new_rows.append(my_row + value)
            return Body(*new_rows)


class Table:
    """
      Represents a table structure.

      Args:
          stream (str): The input stream containing the table structure.
          prepared (bool, optional): Indicates if the stream is already prepared. Default is False.

      Raises:
          TableError: If the provided table structure is incorrect.

      Attributes:
          header (Header): The header of the table.
          body (Body): The body of the table.

      Properties:
          header (Header): The getter and setter for the table's header.
          body (Body): The getter and setter for the table's body.

      Methods:
          column(col: int | str):
              Retrieve a full column of the table.

          row(row: int):
              Retrieve a specific row of the table.

          cells(row: int, column: int) -> Cell:
              Retrieve a specific cell of the table.

      Magic Methods:
          __getitem__(item: int | str):
              Retrieve a specific row or column of the table.

          __repr__():
              Return a string representation of the table.
      """

    __name__ = ''

    def __init__(self, stream: str = None, prepared: bool = False, structure: Literal['xml', 'txt'] = 'txt', **kwargs):
        """
        Initialize the Table instance.

        Args:
            stream (str): The input stream containing the table structure.
            prepared (bool, optional): Indicates if the stream is already prepared. Default is False.

        Raises:
            TableError: If the provided table structure is incorrect.
        """
        if stream is None:
            if kwargs.get('header'):
                header: Header = kwargs.get('header')
                if isinstance(header, Header):
                    self._header = header
                else:
                    raise TableError('Header must be an instance of Header class')
            else:
                header = Header()
            if kwargs.get('body'):
                body: Body = kwargs.get('body')
                if isinstance(body, Body):
                    for row_ in body:
                        for cell_, column_name in zip(row_, header):
                            cell_.name = column_name
                    self._body = body
                else:
                    raise TableError('Body must be an instance of Body class')
            else:
                body = Body()
            self._header = header
            self._body = body
        else:
            if prepared:
                self._stream = stream
            else:
                self._stream = _prepare_table(stream, structure)
            self._header = Header(_define_header(self._stream, structure))
            self._body = Body(_define_body(self._stream, structure))

    @property
    def header(self) -> Header:
        return self._header

    @header.setter
    def header(self, value: list | Header | str):
        if isinstance(value, list):
            new_header = Header()
            new_header.columns = value
            self._header = new_header
        elif isinstance(value, str):
            self._header = Header(value)
        else:
            try:
                # trying to iterate over the value
                iter(value)
                self.header = list(*value)
            except TypeError:
                self._header = value
        self.df.columns = value

    @property
    def body(self) -> Body:
        return self._body

    @body.setter
    def body(self, value: list[Row] | Body | str):
        if isinstance(value, list):
            new_body = Body()
            new_body.rows = value
            self._body = new_body
        elif isinstance(value, str):
            self._body = Body(value)
        else:
            self._body = value
        for row in self._body:
            for i, cell in enumerate(row):
                cell.name = self._header.columns[i]

    def column(self, col: int | str):
        """
        Retrieve a full column of the table.

        Args:
            col (int | str): The column number or value.

        Returns:
            Column: The entire column.
        """
        if isinstance(col, str):
            index = self._header.columns.index(col)
            return Column(col) + self._body.columns[index]
        else:
            return Column(self._header[col], *self._body.columns[col])

    def row(self, row: int):
        """
        Retrieve a specific row of the table.

        Args:
            row (int): The row number.

        Returns:
            Row: The requested row.
        """
        if row == 0:
            return Row(*self._header)
        return self.__getitem__(row)

    def cells(self, row: int, column: int) -> Cell:
        """
        Retrieve a specific cell of the table.

        Args:
            row (int): The row index.
            column (int): The column index.
        Returns:
            Cell: The requested cell.
        """
        return self[row][column]

    @property
    def df(self):
        data = [[c.value for c in row] for row in self._body.rows]
        df = pd.DataFrame(data, columns=self._header.columns)
        # cast data types
        return df

    @property
    def ws(self):
        """
        Return a worksheet representation of the table.

        Returns:
            openpyxl.Worksheet: The worksheet representation of the table.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self._header.columns)
        for row in self._body:
            ws.append(row.to_list())
        return ws

    def drop_column(self, column: str | int):
        if isinstance(column, str):
            for row in self._body:
                row.cells.remove(row.cells[self._header.columns.index(column)])
            self._header.columns.remove(column)
        elif isinstance(column, int):
            for row in self._body:
                row.cells.remove(row.cells[column])
            self._header.columns.remove(self._header.columns[column])
        return self

    def drop_row(self, row: int):
        self._body.rows.remove(self._body.rows[row])
        return self

    def drop_rows(self, rows: list[int]):
        for row in rows:
            self.drop_row(row)
        return self

    def drop_rows_if(self, condition: Callable[[Row], bool]):
        for i, row in self._body.rows:
            if condition(row):
                self.drop_row(i)
        return self

    def drop_columns_if(self, condition: Callable[[Column], bool]):
        for i, column in self._body.columns:
            if condition(column):
                self.drop_column(i)
        return self

    def drop_columns(self, columns: list[str]):
        for column in columns:
            self.drop_column(column)
        return self

    def sort(self, by: str, ascending: bool = True):
        self._body.rows.sort(key=lambda row: row.cells[self._header.columns.index(by)].value, reverse=not ascending)
        return self

    @classmethod
    def from_ws(cls, worksheet: openpyxl.Workbook.worksheets):
        """
        Create a Table instance from an openpyxl worksheet.

        Args:
            worksheet (openpyxl.Workbook.worksheets): The worksheet to convert into a Table.

        Returns:
            Table: The created Table instance.
        """
        table = cls()
        table.__name__ = worksheet.title
        table.header = [cell.value for cell in worksheet[1]]
        table.body = Body(*[Row(*list(row)) for row in worksheet.iter_rows(min_row=2)])
        return table

    @classmethod
    def from_dataframe(cls, dataframe: pd.DataFrame) -> Table:
        """
        Create a Table instance from a pandas DataFrame.

        Args:
            dataframe (pandas.DataFrame): The DataFrame to convert into a Table.

        Returns:
            Table: The created Table instance.
        """
        table = cls()
        table.header = list(dataframe.columns)
        table.body = Body(*[Row(*list(row)) for row in dataframe.values])
        return table

    @classmethod
    def from_csv(cls, file_path, delimiter=','):
        """
        Create a Table instance from a CSV file.

        Args:
            file_path (str): The path to the CSV file.
            delimiter (str, optional): The delimiter used in the CSV file. Default is ','.

        Returns:
            Table: The created Table instance.
        """

        with open(file_path, 'r') as file:
            reader = csv.reader(file, delimiter=delimiter)
            rows = list(reader)

        table = cls()
        table.header = rows[0]
        table.body = [Row(*row) for row in rows[1:]]

        return table

    @classmethod
    def from_json(cls, file_path):
        """
        Create a Table instance from a JSON file.

        Args:
            file_path (str): The path to the JSON file.

        Returns:
            Table: The created Table instance.
        """
        with open(file_path, 'r') as file:
            data = json.load(file)
        table = cls()
        table.header = data['header']
        table.body = [Row(*row) for row in data['body']]
        return table

    @classmethod
    def from_dict(cls, dictionary: dict):
        """
        Create a Table instance from a dictionary.

        Args:
            dictionary (dict): The dictionary to convert into a Table.

        Returns:
            Table: The created Table instance.
        """
        table = cls()
        table.header = dictionary['header']
        table.body = [Row(*row) for row in dictionary['body']]
        return table

    def from_sql(self, connection: str, table_name: str):
        """
        Create a Table instance from a SQL table.

        Args:
            connection (str): The connection string to the SQL database.
            table_name (str): The name of the table to convert into a Table.

        Returns:
            Table: The created Table instance.
        """
        pass

    def to_dict(self) -> dict:
        """
        Convert the Table instance into a dictionary.
        :return:
        """
        return {'header': self.header, 'body': self.body}

    def to_excel(self, file_path: str, sheet_name: str = None, index: bool = False):
        """
        Save the Table instance to an Excel file.

        Args:
            file_path (str): The path to save the Excel file.
            sheet_name (str, optional): The name of the sheet. Defaults to None.
            index (bool, optional): Whether to include the index. Defaults to False.
        """
        if sheet_name is None:
            if self.__name__ is None:
                sheet_name = 'Sheet1'
            else:
                sheet_name = self.__name__
        sheet_name = re.sub(r'/', '', sheet_name[:31])
        self.df.to_excel(file_path, sheet_name=sheet_name, index=index)

    def insert_query(self, *what_columns: str | Header,
                     into: str, values_per_query: int = 0) -> Generator[str, None, None]:
        if not what_columns:
            what_columns = self.header

        columns_to_insert = []
        for col in what_columns:
            if isinstance(col, str):
                columns_to_insert.append(col)
            elif isinstance(col, Header):
                columns_to_insert.extend(col)

        if values_per_query <= 0:
            values_per_query = len(self.body)

        for chunk_start in range(0, len(self.body), values_per_query):
            chunk = self.body[chunk_start: chunk_start + values_per_query]
            values_list = []
            for row in chunk:
                if len(row) < len(columns_to_insert):
                    for i in range(len(columns_to_insert) - len(row), 0, -1):
                        c_ = Cell('null')
                        c_.name = columns_to_insert[-i]
                        row += c_
                columns_to_insert_ = columns_to_insert[:len(row)]
                values = [
                    f"'{row[col]}'".replace(
                        "'null'", 'null'
                    ) if isinstance(row[col], str) else f'{row[col]}' for col in columns_to_insert_
                ]
                values_list.append(f"({', '.join(values)})")
            query = f"INSERT INTO {into} ({', '.join(columns_to_insert_)}) VALUES "
            values_str = ", ".join(values_list)
            yield query + values_str

        return None  # Generator doesn't need to return anything

    def to_insert(self, *what_columns: str | Header, into: str = None, values_per_query: int = 0) -> str:
        if into is None:
            into = self.__name__
        return ';\n'.join(self.insert_query(*what_columns, into=into, values_per_query=values_per_query)) + ';'

    def update_query(self, *what_columns: str | Header,
                     what_table: str, where: str | Header = 'True') -> Generator[str, None, None]:
        if not what_columns:
            what_columns = self.header

        columns_to_update = []
        for col in what_columns:
            if isinstance(col, str):
                columns_to_update.append(col)
            elif isinstance(col, Header):
                columns_to_update.extend(col)

        query = f"UPDATE {what_table} SET "

        where_clause = None
        where_clause_in_row = True
        if where not in self.header.columns:
            where_clause = where
            where_clause_in_row = False

        for chunk_start in range(0, len(self.body)):
            row = self.body[chunk_start]
            if len(row) < len(columns_to_update):
                for i in range(len(columns_to_update) - len(row), 0, -1):
                    c_ = Cell('null')
                    c_.name = columns_to_update[-i]
                    row += c_
            columns_to_update_ = columns_to_update[:len(row)]
            set_values = [f"{col} = '{row[col]}'".replace(
                "'null'", 'null'
            ) if isinstance(row[col], str) else f"{col} = {row[col]}" for col in
                          columns_to_update_]
            if where_clause_in_row:  # if column passed
                if not row[where]:
                    where_clause = 'true'
                else:
                    where_clause = f"{where} = '{row[where]}'".replace(
                        "'null'", 'null'
                    ) if isinstance(row[where], str) else f"{where} = {row[where]}"
            set_values_str = ", ".join(set_values)
            yield query + set_values_str + " WHERE " + where_clause

        return None  # Generator doesn't need to return anything

    def to_update(self, *what_columns: str | Header,
                  what_table: str = None, where: str | Header = 'True') -> str:
        if what_table is None:
            what_table = self.__name__
        return ';\n'.join(self.update_query(*what_columns, what_table=what_table, where=where)) + ';'

    def __getitem__(self, item: int | str):
        if isinstance(item, int):
            if item == 0:
                return self._header
            elif item < 0:
                return self._body[item]
            return self._body[item - 1]
        elif isinstance(item, str):
            return self.column(self._header[item])
        else:
            return self.column(item)

    def __repr__(self):
        return f'{self.__name__} Table[{len(self._body) + int(bool(self.header))} Rows {len(self._header)} Columns]'

    def __str__(self):
        return f"{self._header.rows}\n " + '\n '.join([str(row) for row in self._body])

    def __len__(self):
        return len(self._body) + int(bool(self.header))

    def __iter__(self):
        for row in self._body:
            yield row

    def __add__(self, other):
        if isinstance(other, Column):
            new_header = self._header + [f'column_{len(self._header) + 1}']
            new_rows = []
            for my_row, value in zip(self, other):
                my_row += value
                new_rows.append(my_row)
            new_body = Body(*new_rows)
            table = Table()
            table.header = new_header
            table.body = new_body
            return table
        elif isinstance(other, Row):
            new_body = self._body + other
            table = Table()
            table.header = self._header
            table.body = new_body
            return table
        elif isinstance(other, Table):
            if len(self) != len(other):
                raise TableError('Tables have different number of rows!')
            new_header = self._header + other.header.columns
            new_body = self._body + other.body.columns
            table = Table()
            table.header = new_header
            table.body = new_body
            return table


class Schema:
    def __init__(self, *tables: Table):
        self._tables = list(tables)

    def __iter__(self):
        for tbl in self._tables:
            yield tbl

    def __add__(self, other):
        if isinstance(other, Table):
            self._tables.append(other)
            return self

    def __getitem__(self, item):
        return self._tables[item]

    def __len__(self):
        return len(self._tables)

    def append(self, table: Table):
        if isinstance(table, Table):
            self._tables.append(table)
        else:
            raise TypeError('Schema can only contain Table objects!')

    def _to_excel_pandas(self, file_path: str, sheet_names: List[str] = None, index: bool = False):
        """
        Save the Schema instance to an Excel file using pandas.
        :param file_path:
        :param sheet_names:
        :param index:
        :return:
        """
        with pd.ExcelWriter(file_path) as writer:
            for table, sheet_name in zip(self._tables, sheet_names):
                table.to_excel(writer, sheet_name=sheet_name, index=index)

    def _to_excel_openpyxl(self, file_path: str, sheet_names: List[str] = None):
        """
        Save the Schema instance to an Excel file using openpyxl.
        :param file_path:
        :param sheet_names:
        :return:
        """
        wb = openpyxl.Workbook()
        for table, sheet_name in zip(self._tables, sheet_names):
            ws = wb.create_sheet(sheet_name)
            for i, row in enumerate(table):
                for j, value in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1, value=value)
        wb.save(file_path)

    def to_excel(self,
                 file_path: str,
                 method: Literal['pandas', 'openpyxl'] = 'pandas',
                 sheet_names: List[str] = None,
                 index: bool = False):
        """
        Save the Schema instance to an Excel file.

        Args:
            file_path (str): The path to save the Excel file.
            method (Literal['pandas', 'openpyxl'], optional): The method to use to save the Excel file. Defaults to 'openpyxl'.
            sheet_names (List[str], optional): The names of the sheets. Defaults to None.
            index (bool, optional): Whether to include the index. Defaults to False.
        """

        def create_names():
            table_names = [table.__name__ for table in self._tables]
            tables_names_dict = {name: table_names.count(name) for name in set(table_names)}
            for i, name in enumerate(table_names):
                if name is None:
                    yield f'Sheet{i}'
                name = str(name)
                if tables_names_dict.get(name, 1) > 1:
                    name += '_1'
                    yield name
                yield name

        if sheet_names is None:
            sheet_names = [_ for _ in create_names()]
        if method == 'pandas':
            self._to_excel_pandas(file_path, sheet_names, index)
        elif method == 'openpyxl':
            self._to_excel_openpyxl(file_path, sheet_names)

    @classmethod
    def from_excel(cls, file_path, sheet_names=None):
        """
        Create a Schema instance from an Excel file.

        Args:
            file_path (str): The path to the Excel file.
            sheet_names (List[str], optional): A list of sheet names to include in the schema.
                If None, all sheets in the Excel file will be included. Default is None.

        Returns:
            Schema: The created Schema instance.
        """
        tables = []
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            table = Table.from_ws(wb[sheet_name])
            tables.append(table)
        return cls(*tables)

        # if sheet_names is None:
        #     sheet_names = pd.read_excel(file_path, None).keys()
        #
        # tables = []
        # for sheet_name in sheet_names:
        #     dataframe = pd.read_excel(file_path, sheet_name=sheet_name)
        #     # dataframe = dataframe.fillna(None)
        #     table = Table.from_dataframe(dataframe)
        #     table.__name__ = sheet_name
        #     tables.append(table)
        #
        # return cls(*tables)

    @classmethod
    def from_csv(cls, file_path, delimiter=','):
        pass

    @classmethod
    def from_sql(cls, executor, table_names=None):
        pass

    @classmethod
    def from_json(cls, file_path):
        pass

    @classmethod
    def from_io(cls, stream, type_: Literal['excel'] = 'excel', **kwargs):
        if type_ == 'excel':
            xlsx = io.BytesIO(stream)
            wb = openpyxl.load_workbook(xlsx)
            tables = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = Table.from_ws(ws)
                tables.append(table)
            return cls(*tables)
        return cls()

    def __str__(self):
        if len(self._tables) == 1:
            return "Schema with 1 table"
        if len(self._tables) == 0:
            return "Empty schema"
        return f"Schema with {len(self._tables)} tables"

    def __repr__(self):
        return f"Schema({', '.join(table.__name__ for table in self._tables)})"


def parse_table(table: str) -> Table:
    table = _prepare_table(table)
    if _is_table(table):
        return Table(table)
    else:
        raise TableError('Wrong table passed!')


def add_row(table: Table, *rows: Row) -> Table:
    for row in rows:
        table.body += row
    return table


def add_column(table: Table, *columns: Column) -> Table:
    for column in columns:
        table.header += column[0]
        table.body += column[1:]
    return table


def dataframe_from_table(table: Table) -> pd.DataFrame:
    """
    Convert a Table instance to a pandas DataFrame.
    :param table:
    :return:
    """
    return pd.DataFrame(table.body.rows, columns=table.header.columns)


def from_dataframe(df: pd.DataFrame) -> Table:
    """
    Convert a pandas DataFrame to a Table instance
    :param df:
    :return: Table instance
    """
    table = Table()
    table.header = list(df.columns)
    rows = []
    for row in df.iterrows():
        rows.append(Row(*list(map(lambda cell: Cell(cell), row[1]))))
    table.body = Body(*rows)
    return table


# todo write tests
if __name__ == '__main__':
    header = Header('1', '2', '3')
    body = Body(*[Row(1, 2, 3), Row(4, 5, 6)])
    table = Table(header=header, body=body)
    q = table.to_insert(into='public.table')
    print(q)
    stop = 'here'
